"""
Excel Files:

Extension:
(xls) or (xlsx)
Workbook: A Group of worksheets
Worksheet: An Excel sheet containing data.

"""
import openpyxl
from openpyxl import workbook
from openpyxl.workbook import Workbook

# Writing Content to Excel Files
"""
wb = workbook.Workbook()
ws = wb.create_sheet()

ws["A1"] = "Value1"
ws.cell(row=1, column=2).value = "Value2"
wb.save("./demo.xlsx")
"""


# Reading Content from Excel File
wb = openpyxl.load_workbook("./demo.xlsx")
# ws = wb.active

ws = wb.get_sheet_by_name(wb.sheetnames[1])
cell1_content = ws.cell(row=1, column=2).value
print(cell1_content)


import pandas as pd

df = pd.read_excel("./demo.xlsx", sheet_name="Sheet1")
print(df.head())


import pandas as pd

data = {'Name': ['Alice', 'Bob'], 'Age': [25, 30]}
df = pd.DataFrame(data)

df.to_excel('output.xlsx', index=False) # index=False prevents writing the DataFrame index to Excel